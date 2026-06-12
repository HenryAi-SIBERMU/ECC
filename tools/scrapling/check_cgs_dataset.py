"""
Check CGS Nickel Smelter Dataset to see if it has:
1. Capacity data (ton/year)
2. Investment data (USD/IDR)
"""
import pandas as pd
import openpyxl

print("="*100)
print("CGS NICKEL SMELTER DATASET - DETAILED ANALYSIS")
print("="*100)

cgs_file = '../../data/raw/ESDM/CGS_Nickel_Smelter_Dataset_V1.xlsx'

# First, check all sheet names
try:
    wb = openpyxl.load_workbook(cgs_file, read_only=True)
    print(f"\n📋 Available sheets in CGS dataset:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet_name}")
    wb.close()
    
    # Read the main sheet (usually first one)
    print(f"\n" + "="*100)
    print(f"READING MAIN SHEET: {wb.sheetnames[0]}")
    print("="*100)
    
    df = pd.read_excel(cgs_file, sheet_name=0)
    
    print(f"\n✅ Successfully loaded CGS dataset")
    print(f"Total rows: {len(df)}")
    print(f"Total columns: {len(df.columns)}")
    
    # Show all columns
    print(f"\n" + "="*100)
    print("ALL COLUMNS IN DATASET:")
    print("="*100)
    for i, col in enumerate(df.columns, 1):
        print(f"{i:3d}. {col}")
    
    # Check for capacity-related columns
    print(f"\n" + "="*100)
    print("🔍 CAPACITY DATA CHECK")
    print("="*100)
    capacity_keywords = ['capacity', 'kapasitas', 'production', 'produksi', 'ton', 'tni', 'output']
    capacity_cols = [col for col in df.columns if any(kw in col.lower() for kw in capacity_keywords)]
    
    if capacity_cols:
        print(f"✅ FOUND {len(capacity_cols)} capacity-related columns:")
        for col in capacity_cols:
            non_null = df[col].notna().sum()
            pct = (non_null / len(df) * 100)
            print(f"\n  📊 Column: {col}")
            print(f"     - Non-null entries: {non_null}/{len(df)} ({pct:.1f}%)")
            print(f"     - Data type: {df[col].dtype}")
            if df[col].dtype in ['float64', 'int64']:
                print(f"     - Min: {df[col].min()}")
                print(f"     - Max: {df[col].max()}")
                print(f"     - Mean: {df[col].mean():.2f}")
            print(f"     - Sample values: {df[col].dropna().head(3).tolist()}")
    else:
        print("❌ No capacity columns found!")
    
    # Check for investment-related columns
    print(f"\n" + "="*100)
    print("💰 INVESTMENT DATA CHECK")
    print("="*100)
    investment_keywords = ['invest', 'nilai', 'cost', 'biaya', 'usd', 'idr', 'dollar', 'rupiah', 'capex']
    investment_cols = [col for col in df.columns if any(kw in col.lower() for kw in investment_keywords)]
    
    if investment_cols:
        print(f"✅ FOUND {len(investment_cols)} investment-related columns:")
        for col in investment_cols:
            non_null = df[col].notna().sum()
            pct = (non_null / len(df) * 100)
            print(f"\n  💵 Column: {col}")
            print(f"     - Non-null entries: {non_null}/{len(df)} ({pct:.1f}%)")
            print(f"     - Data type: {df[col].dtype}")
            if df[col].dtype in ['float64', 'int64']:
                print(f"     - Min: {df[col].min()}")
                print(f"     - Max: {df[col].max()}")
                print(f"     - Mean: {df[col].mean():.2f}")
            print(f"     - Sample values: {df[col].dropna().head(3).tolist()}")
    else:
        print("❌ No investment columns found!")
    
    # Check for company name and location columns
    print(f"\n" + "="*100)
    print("📍 COMPANY & LOCATION DATA CHECK")
    print("="*100)
    name_keywords = ['name', 'nama', 'company', 'perusahaan', 'smelter', 'facility']
    name_cols = [col for col in df.columns if any(kw in col.lower() for kw in name_keywords)]
    
    if name_cols:
        print(f"✅ FOUND {len(name_cols)} name-related columns:")
        for col in name_cols:
            print(f"\n  🏢 Column: {col}")
            print(f"     - Sample values:")
            for val in df[col].dropna().head(5):
                print(f"       • {val}")
    
    location_keywords = ['location', 'lokasi', 'province', 'provinsi', 'region', 'area', 'sulawesi']
    location_cols = [col for col in df.columns if any(kw in col.lower() for kw in location_keywords)]
    
    if location_cols:
        print(f"\n✅ FOUND {len(location_cols)} location-related columns:")
        for col in location_cols:
            print(f"\n  📍 Column: {col}")
            if df[col].dtype == 'object':
                print(f"     - Unique values: {df[col].nunique()}")
                print(f"     - Sample values:")
                for val in df[col].dropna().value_counts().head(5).index:
                    count = (df[col] == val).sum()
                    print(f"       • {val} ({count} entries)")
    
    # Sulawesi-specific analysis
    print(f"\n" + "="*100)
    print("🏝️ SULAWESI SMELTERS CHECK")
    print("="*100)
    if location_cols:
        for col in location_cols:
            sulawesi_mask = df[col].str.contains('Sulawesi', case=False, na=False)
            sulawesi_count = sulawesi_mask.sum()
            if sulawesi_count > 0:
                print(f"\n✅ Found {sulawesi_count} Sulawesi entries in column '{col}'")
                print(f"   Locations:")
                for loc in df[sulawesi_mask][col].value_counts().head(10).index:
                    count = (df[sulawesi_mask][col] == loc).sum()
                    print(f"   • {loc}: {count} smelters")
                
                # Show Sulawesi smelters with capacity if available
                if capacity_cols:
                    print(f"\n   📊 Sulawesi smelters with capacity:")
                    sulawesi_df = df[sulawesi_mask]
                    for idx, row in sulawesi_df.head(10).iterrows():
                        name = row[name_cols[0]] if name_cols else f"Smelter #{idx}"
                        cap_val = row[capacity_cols[0]] if capacity_cols else "N/A"
                        print(f"   • {name}: {cap_val}")
    
    # Show sample full rows
    print(f"\n" + "="*100)
    print("📋 SAMPLE DATA (First 3 Rows)")
    print("="*100)
    print(df.head(3).to_string())
    
    # Summary & Recommendation
    print(f"\n" + "="*100)
    print("💡 SUMMARY & RECOMMENDATION")
    print("="*100)
    
    has_capacity = len(capacity_cols) > 0
    has_investment = len(investment_cols) > 0
    has_location = len(location_cols) > 0
    has_name = len(name_cols) > 0
    
    print(f"\n✅ Data Availability:")
    print(f"  - Capacity data: {'✅ YES' if has_capacity else '❌ NO'}")
    print(f"  - Investment data: {'✅ YES' if has_investment else '❌ NO'}")
    print(f"  - Company names: {'✅ YES' if has_name else '❌ NO'}")
    print(f"  - Location data: {'✅ YES' if has_location else '❌ NO'}")
    
    if has_capacity and has_name and has_location:
        print(f"\n✅ RECOMMENDATION: CGS dataset BISA DIPAKAI untuk merge dengan MinerbaOne")
        print(f"   - Match by: Company name + Location")
        print(f"   - Add capacity data to MinerbaOne permits")
        if has_investment:
            print(f"   - Add investment data to MinerbaOne permits")
        else:
            print(f"   - Investment: Use BPS PMDN data as alternative")
    else:
        print(f"\n⚠️ RECOMMENDATION: CGS dataset kurang lengkap")
        print(f"   - Consider alternative sources for missing fields")
    
    # Export to CSV for easier inspection
    output_file = 'output/cgs_dataset_extracted.csv'
    df.to_csv(output_file, index=False)
    print(f"\n💾 CGS data exported to: {output_file}")
    
except FileNotFoundError:
    print(f"\n❌ CGS dataset not found at: {cgs_file}")
    print(f"   Please check file path!")
except Exception as e:
    print(f"\n❌ Error reading CGS dataset: {str(e)}")
    import traceback
    traceback.print_exc()

print("\n" + "="*100)
print("ANALYSIS COMPLETE")
print("="*100)
