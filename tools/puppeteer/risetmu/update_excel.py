import json
import os
from openpyxl import load_workbook

file_path = r'k:\Shared drives\LP2M\1. Agenda\2026\5. Data Lokus\Format Data Lokus LPPM SIBERMU 2026.xlsx'
json_path = r'C:\Users\yooma\OneDrive\Desktop\duniahub\client\4. Celios2\tools\puppeteer\risetmu\links.json'

if not os.path.exists(json_path):
    print("links.json not found! Please run the puppeteer script first.")
    exit(1)

with open(json_path, 'r', encoding='utf-8') as f:
    links = json.load(f)

wb = load_workbook(file_path)
ws = wb['Data Lokus']

updated = 0
for row_num in range(5, 35):
    nama_cell = ws.cell(row=row_num, column=2).value
    if nama_cell and isinstance(nama_cell, str):
        for name_key, url in links.items():
            # Check if the scraped name matches part of the cell name
            if name_key.lower().split()[0] in nama_cell.lower():
                cell = ws.cell(row=row_num, column=9)
                cell.value = f'=HYPERLINK("{url}", "Lihat Kontrak")'
                cell.hyperlink = None
                cell.style = "Hyperlink"
                updated += 1
                print(f"Updated row {row_num} for {name_key}")
                break

wb.save(file_path)
print(f"Done! Updated {updated} hyperlinks.")
