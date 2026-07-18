import os
import csv
from openpyxl import load_workbook

def main():
    print("Mengekstraksi Dataset Historis GFN Indonesia (1 Dekade)...")
    
    file_path = 'data/raw/gfn/NEFBA_Data.xlsx'
    if not os.path.exists(file_path):
        print(f"File {file_path} tidak ditemukan.")
        return

    # Gunakan read_only=True untuk menghemat RAM
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb['national_data']
    
    header = []
    indonesia_rows = []
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            idx_iso3 = header.index('country_ISO3')
            idx_year = header.index('year')
        else:
            if row[idx_iso3] == 'IDN':
                # Filter 1 dekade (2014 - 2023)
                # Kita skip 2024 dan 2025 karena biasanya data PBB lag 1-2 tahun (berisi 0)
                year = row[idx_year]
                if isinstance(year, int) and 2014 <= year <= 2023:
                    indonesia_rows.append(row)
                
    wb.close()
    
    # Sort descending (tahun terbaru di atas)
    indonesia_rows.sort(key=lambda x: x[idx_year], reverse=True)

    out_dir = 'data/processed'
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'nasional_gfn_historis_1_dekade.csv')
    
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(indonesia_rows)
        
    print(f"Berhasil! Data diekstrak ke: {out_path}")
    print(f"Total baris diekstrak: {len(indonesia_rows)} tahun (2014-2023).")

if __name__ == '__main__':
    main()
