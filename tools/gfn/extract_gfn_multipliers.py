import os
import csv
from openpyxl import load_workbook

def main():
    print("Mengekstraksi Faktor Multiplier GFN untuk Indonesia (Low Memory Mode)...")
    
    file_path = 'data/raw/gfn/NEFBA_Data.xlsx'
    if not os.path.exists(file_path):
        print(f"File {file_path} tidak ditemukan.")
        return

    # Gunakan read_only=True untuk menghemat RAM
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb['national_data']
    
    header = []
    indonesia_rows = []
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            # Cari indeks kolom yang dibutuhkan
            idx_iso3 = header.index('country_ISO3')
            idx_year = header.index('year')
            idx_area_forest = header.index('area_forest_ha')
            idx_biocap_forest = header.index('biocap_forest_gha')
            idx_area_crop = header.index('area_crop_ha')
            idx_biocap_crop = header.index('biocap_crop_gha')
            idx_area_built = header.index('area_built_ha')
            idx_biocap_built = header.index('biocap_built_gha')
        else:
            if row[idx_iso3] == 'IDN':
                indonesia_rows.append(row)
                
    wb.close()
    
    # Ambil 5 tahun terakhir
    indonesia_rows.sort(key=lambda x: x[idx_year], reverse=True)
    indonesia_rows = indonesia_rows[:5]
    
    results = []
    for row in indonesia_rows:
        year = row[idx_year]
        
        # Forest
        area_forest = row[idx_area_forest] or 0
        biocap_forest = row[idx_biocap_forest] or 0
        forest_mult = biocap_forest / area_forest if area_forest > 0 else 0
            
        # Cropland
        area_crop = row[idx_area_crop] or 0
        biocap_crop = row[idx_biocap_crop] or 0
        crop_mult = biocap_crop / area_crop if area_crop > 0 else 0
            
        # Built-up land
        area_built = row[idx_area_built] or 0
        biocap_built = row[idx_biocap_built] or 0
        built_mult = biocap_built / area_built if area_built > 0 else 0
            
        results.append({
            'Negara': 'Indonesia',
            'Tahun': year,
            'Tipe_Lahan': 'Hutan (Forest)',
            'Indikator': 'Biocapacity Multiplier (gha/ha)',
            'Nilai_Faktor': round(forest_mult, 4),
            'Deskripsi': 'Mengubah 1 Hektar deforestasi hutan menjadi Global Hectares (gha)'
        })
        
        results.append({
            'Negara': 'Indonesia',
            'Tahun': year,
            'Tipe_Lahan': 'Pertanian (Cropland)',
            'Indikator': 'Biocapacity Multiplier (gha/ha)',
            'Nilai_Faktor': round(crop_mult, 4),
            'Deskripsi': 'Mengubah 1 Hektar alih fungsi lahan pertanian menjadi gha'
        })
        
        results.append({
            'Negara': 'Indonesia',
            'Tahun': year,
            'Tipe_Lahan': 'Lahan Terbangun (Built-up)',
            'Indikator': 'Biocapacity Multiplier (gha/ha)',
            'Nilai_Faktor': round(built_mult, 4),
            'Deskripsi': 'Mengubah 1 Hektar area terbangun/tambang menjadi gha'
        })

    # Simpan ke CSV manual (tanpa pandas)
    out_dir = 'data/processed'
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'nasional_konversi_gfn.csv')
    
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['Negara', 'Tahun', 'Tipe_Lahan', 'Indikator', 'Nilai_Faktor', 'Deskripsi'])
        writer.writeheader()
        writer.writerows(results)
        
    print(f"Berhasil! Data diekstrak ke: {out_path}")
    for r in results[:6]:
        print(r)

if __name__ == '__main__':
    main()
